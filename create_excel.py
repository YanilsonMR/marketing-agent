"""Ejecuta este script una vez para crear leads.xlsx con datos de ejemplo."""

from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()

# --- Hoja de leads ---
leads_ws = wb.active
leads_ws.title = "leads"

leads_headers = [
    "ID", "Title", "Contact Name", "Title Contact", "Company Name",
    "Email", "Email Status", "Company Country", "Company State", "Size",
    "Hiring Profile 1", "Hiring Profile 2", "Serie", "Industry",
    "ICP Rank", "Tier",
]

for col, header in enumerate(leads_headers, 1):
    cell = leads_ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)

sample_leads = [
    [
        "001", "CTO", "Maria Lopez", "CTO", "NovaTech Solutions",
        "maria@novatech.io", "", "USA", "California", "85",
        "ML Engineer", "Data Scientist", "Series A", "Fintech",
        "1", "3",
    ],
    [
        "002", "Founder", "James Chen", "Founder & CEO", "DataPulse",
        "james@datapulse.com", "", "USA", "New York", "120",
        "AI Engineer", "MLOps Engineer", "Series B", "HealthTech",
        "1", "3",
    ],
    [
        "003", "HR Director", "Sarah Kim", "VP People", "CloudScale",
        "sarah@cloudscale.dev", "", "Canada", "Ontario", "200",
        "Backend Engineer", "DevOps Engineer", "Series C", "SaaS",
        "2", "2",
    ],
    [
        "004", "VP Engineering", "Carlos Rivera", "VP Engineering", "GreenGrid",
        "carlos@greengrid.co", "", "USA", "Texas", "45",
        "Full Stack Engineer", "Data Engineer", "Seed", "CleanTech",
        "3", "1",
    ],
    [
        "005", "CEO", "Anna Petrov", "CEO", "QuantumLeap AI",
        "anna@quantumleap.ai", "", "USA", "Massachusetts", "30",
        "AI Researcher", "ML Engineer", "Series A", "AI/ML",
        "1", "3",
    ],
]

for row_data in sample_leads:
    leads_ws.append(row_data)

for col in leads_ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    leads_ws.column_dimensions[col[0].column_letter].width = max_len + 3

# --- Hoja de output ---
output_ws = wb.create_sheet("output")

output_headers = [
    "ID", "Contact Name", "Email Subject", "Email Body",
    "Generated At", "Approved", "Sent",
]

for col, header in enumerate(output_headers, 1):
    cell = output_ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)

for col in output_ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    output_ws.column_dimensions[col[0].column_letter].width = max_len + 3

wb.save("leads.xlsx")
print("leads.xlsx creado con las hojas 'leads' y 'output'.")
print("5 leads de ejemplo incluidos. Editalos o agrega los tuyos.")
