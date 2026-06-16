"""
Excel Service — reads leads and manages output data in an Excel workbook.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from config.settings import Settings


COLUMN_MAP = {
    "id":               "ID",
    "title":            "Title",
    "contact_name":     "Contact Name",
    "title_contact":    "Title Contact",
    "company_name":     "Company Name",
    "email":            "Email",
    "email_status":     "Email Status",
    "country":          "Company Country",
    "state":            "Company State",
    "size":             "Size",
    "hiring_profile_1": "Hiring Profile 1",
    "hiring_profile_2": "Hiring Profile 2",
    "serie":            "Serie",
    "industry":         "Industry",
    "icp_rank":         "ICP Rank",
    "tier":             "Tier",
}

REVERSE_MAP = {v: k for k, v in COLUMN_MAP.items()}


class ExcelService:
    """Reads leads and manages output data in an Excel workbook."""

    def __init__(self, settings: Settings) -> None:
        self._excel_file = settings.excel_file

    def _get_workbook(self):
        """Load and return the Excel workbook."""
        path = Path(self._excel_file)
        if not path.exists():
            raise FileNotFoundError(
                f"Archivo Excel '{self._excel_file}' no encontrado. "
                "Ejecuta: python create_excel.py"
            )
        return load_workbook(self._excel_file)

    def _get_headers(self, ws) -> list[str]:
        """Read header row from a worksheet."""
        return [cell.value for cell in ws[1]]

    def _rows_as_dicts(self, ws) -> list[dict]:
        """Convert worksheet rows to list of dicts using header row as keys."""
        headers = self._get_headers(ws)
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            rows.append({
                headers[i]: (row[i] if row[i] is not None else "")
                for i in range(len(headers))
            })
        return rows

    def read_leads(self, tier: str = None, icp_rank: str = None) -> list[dict]:
        """Read all rows from 'leads' sheet, optionally filtering."""
        wb = self._get_workbook()
        ws = wb["leads"]
        records = self._rows_as_dicts(ws)
        wb.close()

        leads = []
        for row in records:
            lead = {}
            for sheet_col, dict_key in REVERSE_MAP.items():
                lead[dict_key] = row.get(sheet_col, "")
            leads.append(lead)

        if tier is not None:
            leads = [l for l in leads if str(l.get("tier", "")).strip() == str(tier).strip()]
        if icp_rank is not None:
            leads = [l for l in leads if str(l.get("icp_rank", "")) == str(icp_rank)]

        return leads

    def get_output_ids(self) -> set[str]:
        """Return set of lead IDs that already have output generated."""
        wb = self._get_workbook()
        ws = wb["output"]
        records = self._rows_as_dicts(ws)
        wb.close()
        return {str(r.get("ID", "")) for r in records if r.get("ID", "")}

    def get_pending_outputs(self) -> list[dict]:
        """Return output rows where Approved = 'pending'."""
        wb = self._get_workbook()
        ws = wb["output"]
        records = self._rows_as_dicts(ws)
        wb.close()
        return [r for r in records if str(r.get("Approved", "")).lower() == "pending"]

    def get_approved_unsent(self) -> list[dict]:
        """Return output rows where Approved = 'yes' and Sent = 'no'."""
        wb = self._get_workbook()
        ws = wb["output"]
        records = self._rows_as_dicts(ws)
        wb.close()
        return [
            r for r in records
            if str(r.get("Approved", "")).lower() == "yes"
            and str(r.get("Sent", "")).lower() == "no"
        ]

    def get_stats(self) -> dict:
        """Return counts for dashboard stats."""
        wb = self._get_workbook()

        leads_records = self._rows_as_dicts(wb["leads"])
        output_records = self._rows_as_dicts(wb["output"])
        wb.close()

        total_leads = len(leads_records)
        total_generated = len(output_records)
        approved = sum(1 for r in output_records if str(r.get("Approved", "")).lower() == "yes")
        rejected = sum(1 for r in output_records if str(r.get("Approved", "")).lower() == "no")
        pending = sum(1 for r in output_records if str(r.get("Approved", "")).lower() == "pending")
        sent = sum(1 for r in output_records if str(r.get("Sent", "")).lower() == "yes")
        unsent = approved - sent

        return {
            "total_leads": total_leads,
            "total_generated": total_generated,
            "approved": approved,
            "rejected": rejected,
            "pending": pending,
            "sent": sent,
            "unsent": unsent,
        }

    def write_output(self, lead: dict, content: dict) -> None:
        """Write a new row to 'output' sheet."""
        wb = self._get_workbook()
        ws = wb["output"]

        row = [
            lead.get("id", ""),
            lead.get("contact_name", ""),
            content.get("email_subject", ""),
            content.get("email_body", ""),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "pending",
            "no",
        ]
        ws.append(row)
        wb.save(self._excel_file)
        wb.close()

    def delete_output(self, lead_id: str) -> bool:
        """Delete a row from 'output' sheet by lead ID. Returns True if found."""
        wb = self._get_workbook()
        ws = wb["output"]
        headers = self._get_headers(ws)
        id_col = headers.index("ID") + 1

        for row in ws.iter_rows(min_row=2):
            if str(row[id_col - 1].value) == str(lead_id):
                ws.delete_rows(row[0].row)
                wb.save(self._excel_file)
                wb.close()
                return True

        wb.close()
        return False

    def update_approval(self, lead_id: str, approved: bool) -> None:
        """Find row in 'output' sheet by ID and update Approved column."""
        wb = self._get_workbook()
        ws = wb["output"]
        headers = self._get_headers(ws)
        id_col = headers.index("ID") + 1
        approved_col = headers.index("Approved") + 1

        for row in ws.iter_rows(min_row=2):
            if str(row[id_col - 1].value) == str(lead_id):
                ws.cell(row=row[0].row, column=approved_col, value="yes" if approved else "no")
                wb.save(self._excel_file)
                wb.close()
                return

        wb.close()

    def update_sent(self, lead_id: str) -> None:
        """Mark lead as sent in 'output' and update Email Status in 'leads'."""
        wb = self._get_workbook()

        # Update output sheet
        output_ws = wb["output"]
        output_headers = self._get_headers(output_ws)
        id_col = output_headers.index("ID") + 1
        sent_col = output_headers.index("Sent") + 1

        for row in output_ws.iter_rows(min_row=2):
            if str(row[id_col - 1].value) == str(lead_id):
                output_ws.cell(row=row[0].row, column=sent_col, value="yes")
                break

        # Update leads sheet
        leads_ws = wb["leads"]
        leads_headers = self._get_headers(leads_ws)
        leads_id_col = leads_headers.index("ID") + 1
        status_col = leads_headers.index("Email Status") + 1

        for row in leads_ws.iter_rows(min_row=2):
            if str(row[leads_id_col - 1].value) == str(lead_id):
                leads_ws.cell(row=row[0].row, column=status_col, value="sent")
                break

        wb.save(self._excel_file)
        wb.close()
