"""
Seed script — genera leads aleatorios en leads.xlsx.

Uso:
    python seed_leads.py          # 20 leads (default)
    python seed_leads.py 50       # 50 leads
    python seed_leads.py 100      # 100 leads
"""

import random
import sys

from openpyxl import Workbook
from openpyxl.styles import Font


# ─── POOLS DE DATOS ───

FIRST_NAMES = [
    "Maria", "James", "Sarah", "Carlos", "Anna", "Miguel", "Elena", "David",
    "Sofia", "Andrew", "Laura", "Daniel", "Valentina", "Robert", "Camila",
    "Thomas", "Isabella", "Kevin", "Natalia", "Marcus", "Diana", "Eric",
    "Paula", "Ryan", "Andrea", "Brian", "Lucia", "Alex", "Carmen", "Jason",
    "Gabriela", "Patrick", "Monica", "Steven", "Adriana", "Chris", "Fernanda",
    "Jonathan", "Catalina", "Peter", "Victoria", "Richard", "Mariana", "Oscar",
    "Juliana", "Sebastian", "Angela", "Nicolas", "Alejandra", "Felipe",
]

LAST_NAMES = [
    "Lopez", "Chen", "Kim", "Rivera", "Petrov", "Garcia", "Martinez", "Smith",
    "Johnson", "Williams", "Brown", "Jones", "Davis", "Miller", "Wilson",
    "Moore", "Taylor", "Anderson", "Jackson", "Harris", "Thompson", "White",
    "Martin", "Lee", "Walker", "Hall", "Allen", "Young", "King", "Wright",
    "Hernandez", "Torres", "Ramirez", "Flores", "Morales", "Cruz", "Reyes",
    "Gutierrez", "Ortiz", "Ramos", "Vargas", "Castillo", "Mendoza", "Santos",
    "Romero", "Navarro", "Dominguez", "Guerrero", "Medina", "Soto",
]

TITLES = [
    ("CEO", "CEO"),
    ("Founder", "Founder & CEO"),
    ("CTO", "CTO"),
    ("VP Engineering", "VP Engineering"),
    ("VP Sales", "VP Sales"),
    ("VP Operations", "VP Operations"),
    ("COO", "COO"),
    ("CFO", "CFO"),
    ("CMO", "CMO"),
    ("HR Director", "Director of People"),
    ("Head of Growth", "Head of Growth"),
    ("Head of Product", "Head of Product"),
    ("Director of Engineering", "Director of Engineering"),
    ("Director of Sales", "Director of Sales"),
    ("Head of Talent", "Head of Talent Acquisition"),
]

COMPANY_PREFIXES = [
    "Nova", "Data", "Cloud", "Green", "Quantum", "Bright", "Blue", "Core",
    "Apex", "Prime", "Pixel", "Flux", "Neo", "Orbit", "Spark", "Vertex",
    "Atlas", "Helix", "Pulse", "Wave", "Stack", "Zen", "Hyper", "Crypto",
    "Aero", "Cyber", "Rapid", "Smart", "Swift", "Deep", "Open", "True",
]

COMPANY_SUFFIXES = [
    "Tech", "Labs", "AI", "Solutions", "Systems", "Software", "IO",
    "Digital", "Cloud", "Analytics", "Works", "Hub", "Logic", "Forge",
    "Scale", "Stack", "Flow", "Pulse", "Mind", "Base", "Grid", "Wave",
]

DOMAINS = [
    ".io", ".com", ".co", ".dev", ".ai", ".tech", ".app", ".net",
]

INDUSTRIES = [
    "Fintech", "HealthTech", "SaaS", "CleanTech", "AI/ML", "EdTech",
    "E-commerce", "Cybersecurity", "PropTech", "InsurTech", "AgriTech",
    "LegalTech", "HRTech", "MarTech", "LogTech", "FoodTech", "BioTech",
    "GovTech", "RetailTech", "TravelTech", "ConstructionTech", "MediaTech",
]

COUNTRIES = [
    ("USA", ["California", "New York", "Texas", "Florida", "Massachusetts",
             "Illinois", "Washington", "Colorado", "Georgia", "Oregon"]),
    ("Canada", ["Ontario", "British Columbia", "Quebec", "Alberta"]),
    ("UK", ["London", "Manchester", "Edinburgh", "Bristol"]),
    ("Germany", ["Berlin", "Munich", "Hamburg", "Frankfurt"]),
    ("Mexico", ["CDMX", "Monterrey", "Guadalajara", "Puebla"]),
    ("Colombia", ["Bogota", "Medellin", "Cali", "Barranquilla"]),
    ("Argentina", ["Buenos Aires", "Cordoba", "Rosario", "Mendoza"]),
    ("Brazil", ["Sao Paulo", "Rio de Janeiro", "Belo Horizonte", "Curitiba"]),
    ("Spain", ["Madrid", "Barcelona", "Valencia", "Sevilla"]),
    ("Chile", ["Santiago", "Valparaiso", "Concepcion"]),
]

SERIES = ["Pre-Seed", "Seed", "Series A", "Series B", "Series C", "Series D"]

HIRING_PROFILES = [
    "ML Engineer", "Data Scientist", "AI Researcher", "Backend Engineer",
    "Frontend Engineer", "Full Stack Engineer", "DevOps Engineer",
    "Data Engineer", "MLOps Engineer", "Cloud Architect", "SRE",
    "Product Manager", "UX Designer", "iOS Developer", "Android Developer",
    "QA Engineer", "Security Engineer", "Blockchain Developer",
    "Embedded Engineer", "Platform Engineer", "Solutions Architect",
    "Technical Writer", "Scrum Master", "Engineering Manager",
    "Sales Engineer", "Customer Success Manager", "Growth Marketer",
    "Account Executive", "SDR", "Business Analyst",
]


# ─── GENERATOR ───

def generate_company_name() -> str:
    """Generate a random company name."""
    prefix = random.choice(COMPANY_PREFIXES)
    suffix = random.choice(COMPANY_SUFFIXES)
    # Avoid combos like "CloudCloud"
    while prefix.lower() in suffix.lower() or suffix.lower() in prefix.lower():
        suffix = random.choice(COMPANY_SUFFIXES)
    return f"{prefix}{suffix}"


def generate_email(first_name: str, company_name: str) -> str:
    """Generate a plausible email address."""
    domain = random.choice(DOMAINS)
    clean_company = company_name.lower().replace(" ", "")
    patterns = [
        f"{first_name.lower()}@{clean_company}{domain}",
        f"{first_name.lower()[0]}@{clean_company}{domain}",
        f"{first_name.lower()}.contact@{clean_company}{domain}",
    ]
    return random.choice(patterns)


def generate_lead(lead_id: int) -> list:
    """Generate a single random lead row."""
    first = random.choice(FIRST_NAMES)
    last = random.choice(LAST_NAMES)
    contact_name = f"{first} {last}"

    title_short, title_full = random.choice(TITLES)
    company = generate_company_name()
    email = generate_email(first, company)

    country, states = random.choice(COUNTRIES)
    state = random.choice(states)

    size = random.choice([
        random.randint(10, 50),
        random.randint(50, 200),
        random.randint(200, 500),
        random.randint(500, 2000),
    ])

    hp1, hp2 = random.sample(HIRING_PROFILES, 2)
    serie = random.choice(SERIES)
    industry = random.choice(INDUSTRIES)

    icp_rank = random.choices([1, 2, 3], weights=[30, 40, 30])[0]

    # Tier distribution: 3 (high priority) more common for rank 1
    if icp_rank == 1:
        tier = random.choices([1, 2, 3], weights=[10, 20, 70])[0]
    elif icp_rank == 2:
        tier = random.choices([1, 2, 3], weights=[20, 40, 40])[0]
    else:
        tier = random.choices([1, 2, 3], weights=[40, 35, 25])[0]

    return [
        str(lead_id).zfill(3),
        title_short,
        contact_name,
        title_full,
        company,
        email,
        "",          # Email Status
        country,
        state,
        str(size),
        hp1,
        hp2,
        serie,
        industry,
        str(icp_rank),
        str(tier),
    ]


def main():
    count = 20
    if len(sys.argv) > 1:
        try:
            count = max(1, int(sys.argv[1]))
        except ValueError:
            print(f"Uso: python seed_leads.py [cantidad]")
            return

    wb = Workbook()

    # --- Hoja de leads ---
    leads_ws = wb.active
    leads_ws.title = "leads"

    headers = [
        "ID", "Title", "Contact Name", "Title Contact", "Company Name",
        "Email", "Email Status", "Company Country", "Company State", "Size",
        "Hiring Profile 1", "Hiring Profile 2", "Serie", "Industry",
        "ICP Rank", "Tier",
    ]

    for col, header in enumerate(headers, 1):
        cell = leads_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    used_names = set()
    for i in range(1, count + 1):
        row = generate_lead(i)
        # Evitar nombres duplicados
        while row[2] in used_names:
            row = generate_lead(i)
        used_names.add(row[2])
        leads_ws.append(row)

    for col in leads_ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        leads_ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)

    # --- Hoja de output ---
    output_ws = wb.create_sheet("output")

    output_headers = [
        "ID", "Contact Name", "Email Subject", "Email Body",
        "Generated At", "Approved", "Sent",
    ]

    for col, header in enumerate(output_headers, 1):
        cell = output_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for col in output_ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        output_ws.column_dimensions[col[0].column_letter].width = max_len + 3

    wb.save("leads.xlsx")

    # Stats
    tiers = {}
    ranks = {}
    for row in leads_ws.iter_rows(min_row=2, values_only=True):
        t = str(row[15])
        r = str(row[14])
        tiers[t] = tiers.get(t, 0) + 1
        ranks[r] = ranks.get(r, 0) + 1

    print(f"leads.xlsx creado con {count} leads aleatorios.")
    print(f"  Tiers:     {', '.join(f'T{k}={v}' for k, v in sorted(tiers.items()))}")
    print(f"  ICP Ranks: {', '.join(f'R{k}={v}' for k, v in sorted(ranks.items()))}")
    print(f"  Hoja 'output' limpia (solo headers).")


if __name__ == "__main__":
    main()
